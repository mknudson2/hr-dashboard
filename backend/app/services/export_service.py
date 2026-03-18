"""Export service for employee data to Excel and PDF formats."""
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch


def export_employees_excel(employees: list, view_mode: str = "standard") -> BytesIO:
    """Export employee data to Excel format.

    Args:
        employees: List of employee dictionaries
        view_mode: "standard" or "compensation"

    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Define headers based on view mode
    if view_mode == "compensation":
        headers = [
            "Employee ID",
            "Name",
            "Type",
            "Department",
            "Team",
            "Cost Center",
            "Base/Hourly Rate",
            "Annualized/Salary",
            "Wage Type",
            "Wage Effective Date",
        ]
    else:
        headers = [
            "Employee ID",
            "Name",
            "Department",
            "Hire Date",
            "Status",
            "Location",
            "Type",
        ]

    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for row_num, emp in enumerate(employees, 2):
        if view_mode == "compensation":
            hourly_rate = emp.get("wage", 0) / 2080 if emp.get("wage") else None
            annual_salary = emp.get("wage")

            row_data = [
                emp.get("employee_id", ""),
                f"{emp.get('first_name', '')} {emp.get('last_name', '')}",
                emp.get("type", ""),
                emp.get("department", ""),
                emp.get("team", ""),
                emp.get("cost_center", ""),
                f"${hourly_rate:,.2f}" if hourly_rate else "N/A",
                f"${annual_salary:,.0f}" if annual_salary else "N/A",
                emp.get("wage_type", ""),
                emp.get("wage_effective_date", ""),
            ]
        else:
            row_data = [
                emp.get("employee_id", ""),
                f"{emp.get('first_name', '')} {emp.get('last_name', '')}",
                emp.get("department", ""),
                emp.get("hire_date", ""),
                emp.get("status", ""),
                emp.get("location", ""),
                emp.get("type", ""),
            ]

        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_employees_pdf(employees: list, view_mode: str = "standard") -> BytesIO:
    """Export employee data to PDF format.

    Args:
        employees: List of employee dictionaries
        view_mode: "standard" or "compensation"

    Returns:
        BytesIO object containing PDF file
    """
    output = BytesIO()

    # Create PDF with landscape orientation for better table fit
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    # Container for PDF elements
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph(
        f"<b>Employee Data Export - {datetime.now().strftime('%Y-%m-%d %H:%M')}</b>",
        styles['Title']
    )
    elements.append(title)
    elements.append(Spacer(1, 0.25 * inch))

    # Prepare table data
    if view_mode == "compensation":
        headers = [
            "ID", "Name", "Type", "Dept", "Team",
            "Cost Center", "Hourly", "Annual", "Wage Type"
        ]

        table_data = [headers]
        for emp in employees:
            hourly_rate = emp.get("wage", 0) / 2080 if emp.get("wage") else None
            annual_salary = emp.get("wage")

            row = [
                emp.get("employee_id", ""),
                f"{emp.get('first_name', '')} {emp.get('last_name', '')}",
                emp.get("type", ""),
                emp.get("department", "")[:15] if emp.get("department") else "",
                emp.get("team", "")[:12] if emp.get("team") else "",
                emp.get("cost_center", ""),
                f"${hourly_rate:,.2f}" if hourly_rate else "N/A",
                f"${annual_salary:,.0f}" if annual_salary else "N/A",
                emp.get("wage_type", ""),
            ]
            table_data.append(row)
    else:
        headers = ["ID", "Name", "Department", "Hire Date", "Status", "Location", "Type"]

        table_data = [headers]
        for emp in employees:
            row = [
                emp.get("employee_id", ""),
                f"{emp.get('first_name', '')} {emp.get('last_name', '')}",
                emp.get("department", "")[:20] if emp.get("department") else "",
                emp.get("hire_date", ""),
                emp.get("status", ""),
                emp.get("location", "")[:15] if emp.get("location") else "",
                emp.get("type", ""),
            ]
            table_data.append(row)

    # Create table
    table = Table(table_data)

    # Style the table
    table_style = TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

        # Body styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ])

    table.setStyle(table_style)
    elements.append(table)

    # Add footer with record count
    elements.append(Spacer(1, 0.25 * inch))
    footer = Paragraph(
        f"<i>Total Records: {len(employees)}</i>",
        styles['Normal']
    )
    elements.append(footer)

    # Build PDF
    doc.build(elements)
    output.seek(0)
    return output
