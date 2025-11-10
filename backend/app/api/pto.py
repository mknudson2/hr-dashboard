from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from typing import Optional, List
from datetime import date, datetime
from pydantic import BaseModel
from app.db import models, database
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter(prefix="/pto", tags=["pto"])


def get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ============================================================================
# PYDANTIC SCHEMAS
# ============================================================================

class PTORecordCreate(BaseModel):
    employee_id: str
    cost_center: str
    pay_period_date: date
    pto_hours: float
    pto_cost: float
    hourly_rate: Optional[float] = None
    notes: Optional[str] = None


class PTORecordUpdate(BaseModel):
    pto_hours: Optional[float] = None
    pto_cost: Optional[float] = None
    hourly_rate: Optional[float] = None
    notes: Optional[str] = None


# ============================================================================
# PTO RECORDS ENDPOINTS
# ============================================================================

@router.get("/records")
def get_pto_records(
    employee_id: Optional[str] = None,
    cost_center: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get PTO records with optional filters."""
    query = db.query(models.PTORecord)

    if employee_id:
        query = query.filter(models.PTORecord.employee_id == employee_id)
    if cost_center:
        query = query.filter(models.PTORecord.cost_center == cost_center)
    if start_date:
        query = query.filter(models.PTORecord.pay_period_date >= start_date)
    if end_date:
        query = query.filter(models.PTORecord.pay_period_date <= end_date)

    records = query.order_by(models.PTORecord.pay_period_date.desc()).all()

    result = []
    for record in records:
        employee = db.query(models.Employee).filter(
            models.Employee.employee_id == record.employee_id
        ).first()

        result.append({
            "id": record.id,
            "employee_id": record.employee_id,
            "employee_name": f"{employee.first_name} {employee.last_name}" if employee else None,
            "cost_center": record.cost_center,
            "pay_period_date": record.pay_period_date.isoformat(),
            "pto_hours": record.pto_hours,
            "pto_cost": record.pto_cost,
            "hourly_rate": record.hourly_rate,
            "notes": record.notes,
        })

    return {"records": result, "total": len(result)}


@router.post("/records")
def create_pto_record(record: PTORecordCreate, db: Session = Depends(get_db)):
    """Create a new PTO record."""
    # Check if record already exists for this employee and pay period
    existing = db.query(models.PTORecord).filter(
        and_(
            models.PTORecord.employee_id == record.employee_id,
            models.PTORecord.pay_period_date == record.pay_period_date
        )
    ).first()

    if existing:
        raise HTTPException(
            status_code=400,
            detail="PTO record already exists for this employee and pay period"
        )

    new_record = models.PTORecord(**record.dict())
    db.add(new_record)
    db.commit()
    db.refresh(new_record)

    return {"message": "PTO record created successfully", "id": new_record.id}


@router.put("/records/{record_id}")
def update_pto_record(
    record_id: int,
    record: PTORecordUpdate,
    db: Session = Depends(get_db)
):
    """Update an existing PTO record."""
    db_record = db.query(models.PTORecord).filter(
        models.PTORecord.id == record_id
    ).first()

    if not db_record:
        raise HTTPException(status_code=404, detail="PTO record not found")

    for key, value in record.dict(exclude_unset=True).items():
        setattr(db_record, key, value)

    db.commit()
    db.refresh(db_record)

    return {"message": "PTO record updated successfully"}


@router.delete("/records/{record_id}")
def delete_pto_record(record_id: int, db: Session = Depends(get_db)):
    """Delete a PTO record."""
    record = db.query(models.PTORecord).filter(
        models.PTORecord.id == record_id
    ).first()

    if not record:
        raise HTTPException(status_code=404, detail="PTO record not found")

    db.delete(record)
    db.commit()

    return {"message": "PTO record deleted successfully"}


# ============================================================================
# SUMMARY AND ANALYTICS ENDPOINTS
# ============================================================================

@router.get("/summary")
def get_pto_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get PTO summary by cost center and overall."""
    query = db.query(models.PTORecord)

    if start_date:
        query = query.filter(models.PTORecord.pay_period_date >= start_date)
    if end_date:
        query = query.filter(models.PTORecord.pay_period_date <= end_date)

    records = query.all()

    # Aggregate by cost center
    cost_center_summary = {}
    total_hours = 0
    total_cost = 0

    for record in records:
        cc = record.cost_center
        if cc not in cost_center_summary:
            cost_center_summary[cc] = {
                "cost_center": cc,
                "total_hours": 0,
                "total_cost": 0,
                "employee_count": set()
            }

        cost_center_summary[cc]["total_hours"] += record.pto_hours
        cost_center_summary[cc]["total_cost"] += record.pto_cost
        cost_center_summary[cc]["employee_count"].add(record.employee_id)
        total_hours += record.pto_hours
        total_cost += record.pto_cost

    # Convert sets to counts
    summary_list = []
    for cc, data in cost_center_summary.items():
        summary_list.append({
            "cost_center": cc,
            "total_hours": round(data["total_hours"], 2),
            "total_cost": round(data["total_cost"], 2),
            "employee_count": len(data["employee_count"])
        })

    return {
        "summary_by_cost_center": summary_list,
        "overall": {
            "total_hours": round(total_hours, 2),
            "total_cost": round(total_cost, 2),
            "cost_center_count": len(cost_center_summary)
        }
    }


# ============================================================================
# EXCEL EXPORT ENDPOINT
# ============================================================================

@router.get("/export")
def export_pto_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Generate Excel report matching the YTD Overtime Report format."""

    # Query PTO records
    query = db.query(models.PTORecord)

    if start_date:
        query = query.filter(models.PTORecord.pay_period_date >= start_date)
    if end_date:
        query = query.filter(models.PTORecord.pay_period_date <= end_date)

    records = query.all()

    if not records:
        raise HTTPException(status_code=404, detail="No PTO records found for the specified date range")

    # Build data structure for Excel
    # Group by cost center and employee
    data_by_cost_center = {}
    pay_periods = sorted(list(set([r.pay_period_date for r in records])))

    for record in records:
        cc = record.cost_center
        emp_id = record.employee_id

        if cc not in data_by_cost_center:
            data_by_cost_center[cc] = {}

        if emp_id not in data_by_cost_center[cc]:
            employee = db.query(models.Employee).filter(
                models.Employee.employee_id == emp_id
            ).first()

            data_by_cost_center[cc][emp_id] = {
                "Employee": f"{employee.last_name}, {employee.first_name}" if employee else emp_id,
                "ID": emp_id,
                "Cost Center": cc,
                **{pd.strftime("%m-%d-%Y"): 0.0 for pd in pay_periods}
            }

        period_key = record.pay_period_date.strftime("%m-%d-%Y")
        data_by_cost_center[cc][emp_id][period_key] = record.pto_cost

    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Create Overview sheet
    overview_ws = wb.create_sheet("Overview")
    overview_headers = ["Cost Center", "Total YTD"] + [pd.strftime("%m-%d-%Y") for pd in pay_periods]
    overview_ws.append(overview_headers)

    # Style overview headers
    for cell in overview_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add overview data
    for cc, employees in data_by_cost_center.items():
        row_data = [cc, 0.0] + [0.0 for _ in pay_periods]

        for emp_data in employees.values():
            for idx, pd in enumerate(pay_periods):
                period_key = pd.strftime("%m-%d-%Y")
                row_data[idx + 2] += emp_data.get(period_key, 0.0)
                row_data[1] += emp_data.get(period_key, 0.0)

        overview_ws.append(row_data)

    # Create individual cost center sheets
    for cc, employees in data_by_cost_center.items():
        ws = wb.create_sheet(cc)

        # Headers
        headers = ["Employee", "ID", "Cost Center"] + [pd.strftime("%m-%d-%Y") for pd in pay_periods] + ["Total_Amount", "Average_Hours", "Employee_Total"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add employee data
        for emp_id, emp_data in employees.items():
            row = [
                emp_data["Employee"],
                emp_data["ID"],
                emp_data["Cost Center"]
            ]

            total_amount = 0.0
            for pd in pay_periods:
                period_key = pd.strftime("%m-%d-%Y")
                value = emp_data.get(period_key, 0.0)
                row.append(value)
                total_amount += value

            # Get actual PTO hours for average calculation
            emp_records = [r for r in records if r.employee_id == emp_id]
            total_hours = sum([r.pto_hours for r in emp_records])
            avg_hours = total_hours / len(pay_periods) if pay_periods else 0

            row.extend([total_amount, round(avg_hours, 6), total_amount])
            ws.append(row)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename
    report_date = datetime.now().strftime("%m-%d-%Y")
    filename = f"YTD PTO Report - {report_date}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================================
# FILE UPLOAD / IMPORT ENDPOINT
# ============================================================================

@router.post("/upload")
async def upload_ot_earnings(
    file: UploadFile = File(...),
    imported_by: str = "System",
    db: Session = Depends(get_db)
):
    """Upload and process OT Earnings Excel file."""

    if not file.filename.endswith(('.xlsx', '.xls', '.XLSX', '.XLS')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")

    try:
        # Read the uploaded file
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents), sheet_name=0)

        # The file has headers in row 2 (index 2)
        # Find the header row
        header_row_idx = None
        for idx, row in df.iterrows():
            if 'Employee' in str(row.values) or 'ID' in str(row.values):
                header_row_idx = idx
                break

        if header_row_idx is None:
            raise HTTPException(status_code=400, detail="Could not find header row in Excel file")

        # Re-read with proper headers
        df = pd.read_excel(BytesIO(contents), sheet_name=0, skiprows=header_row_idx)

        # Rename columns properly
        df.columns = ['Employee', 'Col1', 'ID', 'SSN', 'Location', 'Col5', 'Chk_Date', 'Rate', 'Col8', 'Hours', 'Col10', 'Amount']

        # Clean the data
        df = df[df['ID'].notna()]  # Remove rows without ID
        df = df[df['ID'] != 'ID']  # Remove duplicate header rows
        df = df[df['Chk_Date'].notna()]  # Remove rows without check date

        # Convert data types
        df['ID'] = df['ID'].astype(str).str.strip()
        df['Hours'] = pd.to_numeric(df['Hours'], errors='coerce')
        df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
        df['Rate'] = pd.to_numeric(df['Rate'], errors='coerce')

        # Parse dates
        df['Chk_Date'] = pd.to_datetime(df['Chk_Date'], errors='coerce')

        # Remove rows with invalid data
        df = df.dropna(subset=['ID', 'Chk_Date', 'Hours', 'Amount'])

        # Track import stats
        records_imported = 0
        records_updated = 0
        records_skipped = 0
        errors = []

        min_date = df['Chk_Date'].min()
        max_date = df['Chk_Date'].max()

        for _, row in df.iterrows():
            try:
                employee_id = str(row['ID']).strip()

                # Get employee to find cost center
                employee = db.query(models.Employee).filter(
                    models.Employee.employee_id == employee_id
                ).first()

                if not employee:
                    errors.append(f"Employee ID {employee_id} not found in database")
                    records_skipped += 1
                    continue

                cost_center = employee.cost_center or employee.department or "Unknown"
                pay_period_date = row['Chk_Date'].date()
                pto_hours = float(row['Hours'])
                pto_cost = float(row['Amount'])
                hourly_rate = float(row['Rate']) if pd.notna(row['Rate']) else None

                # Check if record exists
                existing = db.query(models.PTORecord).filter(
                    and_(
                        models.PTORecord.employee_id == employee_id,
                        models.PTORecord.pay_period_date == pay_period_date
                    )
                ).first()

                if existing:
                    # Update existing record
                    existing.pto_hours = pto_hours
                    existing.pto_cost = pto_cost
                    existing.hourly_rate = hourly_rate
                    existing.cost_center = cost_center
                    existing.updated_at = datetime.now()
                    records_updated += 1
                else:
                    # Create new record
                    new_record = models.PTORecord(
                        employee_id=employee_id,
                        cost_center=cost_center,
                        pay_period_date=pay_period_date,
                        pto_hours=pto_hours,
                        pto_cost=pto_cost,
                        hourly_rate=hourly_rate
                    )
                    db.add(new_record)
                    records_imported += 1

            except Exception as e:
                errors.append(f"Error processing row for employee {employee_id}: {str(e)}")
                records_skipped += 1
                continue

        # Commit all records
        db.commit()

        # Create import history record
        import_history = models.PTOImportHistory(
            file_name=file.filename,
            imported_by=imported_by,
            records_imported=records_imported + records_updated,
            start_date=min_date.date() if pd.notna(min_date) else None,
            end_date=max_date.date() if pd.notna(max_date) else None,
            notes=f"Created: {records_imported}, Updated: {records_updated}, Skipped: {records_skipped}",
            status="success" if records_skipped == 0 else "partial_success"
        )
        db.add(import_history)
        db.commit()

        return {
            "message": "File processed successfully",
            "records_created": records_imported,
            "records_updated": records_updated,
            "records_skipped": records_skipped,
            "total_processed": records_imported + records_updated,
            "errors": errors[:10] if errors else [],  # Return first 10 errors
            "date_range": {
                "start": min_date.date().isoformat() if pd.notna(min_date) else None,
                "end": max_date.date().isoformat() if pd.notna(max_date) else None
            }
        }

    except Exception as e:
        # Log error to import history
        import_history = models.PTOImportHistory(
            file_name=file.filename,
            imported_by=imported_by,
            records_imported=0,
            notes=f"Import failed: {str(e)}",
            status="failed"
        )
        db.add(import_history)
        db.commit()

        raise HTTPException(
            status_code=500,
            detail=f"Failed to process file: {str(e)}"
        )


# ============================================================================
# IMPORT HISTORY ENDPOINTS
# ============================================================================

@router.get("/import-history")
def get_import_history(db: Session = Depends(get_db)):
    """Get PTO import history."""
    history = db.query(models.PTOImportHistory).order_by(
        models.PTOImportHistory.import_date.desc()
    ).all()

    return {
        "history": [
            {
                "id": h.id,
                "file_name": h.file_name,
                "import_date": h.import_date.isoformat() if h.import_date else None,
                "imported_by": h.imported_by,
                "records_imported": h.records_imported,
                "start_date": h.start_date.isoformat() if h.start_date else None,
                "end_date": h.end_date.isoformat() if h.end_date else None,
                "status": h.status,
                "notes": h.notes,
            }
            for h in history
        ],
        "total": len(history)
    }
